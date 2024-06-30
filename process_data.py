#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jun  3 11:10:13 2024

@author: vishalr
"""

from openpyxl import load_workbook
import os
import numpy as np
import pandas as pd
import pickle
from tqdm import tqdm
import random

random.seed(42)

##################################
##################################

#change this to the folder where you store your data
data_dir = "mcmc_chromatin_w_nodeID_modified/"

#each of the two data frames below have 20,000 rows, each corresponding to one sample from the original graph
#each sample consists of 21 nodes; node_labels contains the names of these 21 nodes of the form Vxyz
#each node is a DNA fragment of length 500 bases; so Vxyz coveres region [500*xyz, 500*xyz+500)
#adjacency_matrix has the flattended adjacecny matrix for each of these 20,000 samples
#so each row is of dimension 21*21 = 441


#node_labels = np.load(data_dir+'df_chr2R_drosophila_ChIA_Drop_0.1_PASS_20000_MCMC_pivot_sample_node_matrix.npy')

#adjacency_matrix = np.load(data_dir+'df_chr2R_drosophila_ChIA_Drop_0.1_PASS_20000_MCMC_pivot.npy')

worklist_node_labels = []
worklist_adjacency_matrix = []

#importing the rest of the files
# 2L = 0; 2R = 1; 3L = 2, 3R = 3

node_labels_2L = np.load(data_dir+'df_chr2L_drosophila_ChIA_Drop_0.1_PASS_20000_MCMC_pivot_sample_node_matrix.npy')
adjacency_matrix_2L = np.load(data_dir+'df_chr2L_drosophila_ChIA_Drop_0.1_PASS_20000_MCMC_pivot.npy')

worklist_node_labels.append(node_labels_2L)
worklist_adjacency_matrix.append(adjacency_matrix_2L)

node_labels_2R = np.load(data_dir+'df_chr2R_drosophila_ChIA_Drop_0.1_PASS_20000_MCMC_pivot_sample_node_matrix.npy')
adjacency_matrix_2R = np.load(data_dir+'df_chr2R_drosophila_ChIA_Drop_0.1_PASS_20000_MCMC_pivot.npy')

worklist_node_labels.append(node_labels_2R)
worklist_adjacency_matrix.append(adjacency_matrix_2R)

node_labels_3L = np.load(data_dir+'df_chr3L_drosophila_ChIA_Drop_0.1_PASS_20000_MCMC_pivot_sample_node_matrix.npy')
adjacency_matrix_3L = np.load(data_dir+'df_chr3L_drosophila_ChIA_Drop_0.1_PASS_20000_MCMC_pivot.npy')

worklist_node_labels.append(node_labels_3L)
worklist_adjacency_matrix.append(adjacency_matrix_3L)

node_labels_3R = np.load(data_dir+'df_chr3R_drosophila_ChIA_Drop_0.1_PASS_20000_MCMC_pivot_sample_node_matrix.npy')
adjacency_matrix_3R = np.load(data_dir+'df_chr3R_drosophila_ChIA_Drop_0.1_PASS_20000_MCMC_pivot.npy')


worklist_node_labels.append(node_labels_3R)
worklist_adjacency_matrix.append(adjacency_matrix_3R)



rep_A = {}
rep_graph_indicator = []
rep_graph_labels = []
rep_node_labels = []
rep_node_attributes = []
rep_edge_attributes = []
#rep_graph_uid = []


#node labels - each node needs a unique label across all chromosomes
#using the mapping chr2L=0, chr2R=1, etc

#node attributes - use one hot encoded sequence
#check if Vids are from 0 or 1
#TODO

#adjacency matrix and edge attributes and features

graph_counter = 0
vertex_counter = 0

for chr_label, chr_adj in enumerate(worklist_adjacency_matrix):
    #print(chr_label)
    rep_graph_labels.extend([chr_label for itr in range(len(chr_adj))])
    for sub_itr, sub_adj in enumerate(chr_adj):
        #reshape the adjacency matrix
        sub_adj = sub_adj[:-1].reshape(21, 21)
        #print(sub_adj)
        
        #retrive the corresponding node list
        sub_node_list = worklist_node_labels[chr_label][sub_itr]
        
        
        #remove duplicate nodes
        sub_node_list_clean = []
        sub_adj_clean = []
        
        adj_to_remove = []
        for clean_itr, clean_node in enumerate(sub_node_list):
            if(clean_node in sub_node_list_clean):
                #if the node already exists in the list - remove this occurance 
                adj_to_remove.append(clean_itr)
            else:
                #if it doesn't already exist, add it to the node list
                sub_node_list_clean.append(clean_node)
                
        #print(sub_node_list)   
        #print(sub_node_list_clean)
        #print(adj_to_remove)
        
        #delete the duplicated rows and columns from adjacency matrix
        sub_adj_clean = np.delete(sub_adj, adj_to_remove, 0)
        sub_adj_clean = np.delete(sub_adj_clean, adj_to_remove, 1)
        
        #add the node labels from this subgraph to list of nodes
        #rep_node_labels.extend(sub_node_list_clean)
        rep_node_labels.extend([itr+1 for itr in range(len(sub_node_list_clean))])
        
        rep_node_attributes.extend(sub_node_list_clean)
        #subgraph membership of the vertices
        rep_graph_indicator.extend([graph_counter for itr in range(len(sub_node_list_clean))])
        
        #add edges using the correct vertex labels
        for itrx in range(len(sub_node_list_clean)):
            for itry in range(itrx+1, len(sub_node_list_clean)):
                if(sub_adj_clean[itrx][itry] >0):
                    #print(itrx, itry,sub_node_list_clean[itrx], sub_node_list_clean[itry])
                    rep_A[vertex_counter+itrx, vertex_counter+itry] = np.absolute(sub_node_list_clean[itrx] - sub_node_list_clean[itry])
        
        #update the counter for number of nodes already in the list
        vertex_counter+=len(sub_node_list_clean)
        
        #update graph counter
        graph_counter+=1

    

exp_name = 'stamp0603'
gnn_files_loc = '/data/shared/vishal/chromatin_embedding/Research-Graph-Embeddings--JJ/data/'+exp_name+'/raw/'

with open(os.path.join(gnn_files_loc, exp_name+'_A.txt'), 'w') as fw:
    with open(os.path.join(gnn_files_loc, exp_name+'_edge_attributes.txt'), 'w') as fe:
        for line in rep_A.keys():
            # line = list(map(str, line))
            # fw.write(','.join(line) + '\n')
            fw.write(f'{line[0]+1},{line[1]+1}\n')
            fe.write(str(rep_A[line]) + '\n')

with open(os.path.join(gnn_files_loc, exp_name+'_graph_indicator.txt'), 'w') as fw:
    for tmp in rep_graph_indicator:
        fw.write(str(tmp+1) + '\n')

with open(os.path.join(gnn_files_loc, exp_name+'_graph_labels.txt'), 'w') as fw:
    for tmp in rep_graph_labels:
        fw.write(str(tmp) + '\n')
        
with open(os.path.join(gnn_files_loc, exp_name+'_node_labels.txt'), 'w') as fw:
    for tmp in rep_node_labels:
        fw.write(str(tmp) + '\n')

#edge_attributes
#with open(os.path.join(gnn_files_loc, exp_name+'_edge_attributes.txt'), 'w') as fw:
#    for tmp in rep_edge_attributes:
#        fw.write(str(tmp) + '\n')

#node_attributes
with open(os.path.join(gnn_files_loc, exp_name+'_node_attributes.txt'), 'w') as fw:
    for line in rep_node_attributes:
        #line = list(map(str, line))
        fw.write(str(line) + '\n')
        


'''
for uid in tqdm(uid_top_classes_list):
    if uid not in folders:
        continue
    #print(uid)
    rsam_graph_uid.append(valid_count)
    uid_map.append(uid)
    
    valid_count += 1
    uid_path = os.path.join(fpath, uid)
    # load the processed features
    dist = np.load(os.path.join(uid_path, 'distance_matrix.npy'))
    edge_list = np.load(os.path.join(uid_path, 'edge_list.npy'))
    node_features = np.load(os.path.join(uid_path, 'node_features.npy'))
    tmp_residual_labels, tmp_atom_labels = [], []
    with open(os.path.join(uid_path, 'node_identities.npy'), 'r') as fr:
        lines = fr.read().splitlines()
    assert len(lines) == dist.shape[0] and len(lines) == node_features.shape[0]
    for line in lines:
        line = line.split('\t')
        tmp_residual_labels.append(residual_mapping[line[1]])
        tmp_atom_labels.append(atom_mapping[line[2]])
    
    # first add the edges in edge_list to rsam_A
    for m in range(edge_list.shape[0]):
        # do not include self-loops
        if edge_list[m][0] == edge_list[m][1]:
            continue
        tmp_edge = (num_nodes + edge_list[m][0], num_nodes + edge_list[m][1])
        tmp_edge_rev = (num_nodes + edge_list[m][1], num_nodes + edge_list[m][0])
        assert dist[edge_list[m][0]][edge_list[m][1]] == dist[edge_list[m][1]][edge_list[m][0]]
        rsam_A[tmp_edge] = 1 / dist[edge_list[m][0]][edge_list[m][1]] ** 2
        rsam_A[tmp_edge_rev] = 1 / dist[edge_list[m][0]][edge_list[m][1]] ** 2

    rsam_graph_indicator += [num_graphs] * dist.shape[0]
    rsam_graph_labels.append(int(uid_top_classes[uid]))
    
    rsam_node_labels += tmp_atom_labels
    
    for n in range(dist.shape[0]):
        one_hot_residual_feature = np.zeros(len(residual_set))
        one_hot_residual_feature[tmp_residual_labels[n]] = 1.0
        # one_hot_atom_feature = np.zeros(len(atom_set))
        # one_hot_atom_feature[tmp_atom_labels[n]] = 1.0
        rsam_node_attributes.append(np.concatenate((node_features[n], one_hot_residual_feature)))

    # update the number of graphs and num_nodes
    num_nodes += dist.shape[0]
    num_graphs += 1

# separate rsam_A out as A and edge_attributes
rsam_A_edge_list = []
rsam_edge_attributes = []
for edge in rsam_A:
    rsam_A_edge_list.append([edge[0], edge[1]])
    rsam_edge_attributes.append(rsam_A[edge])

print('number of edges', len(rsam_A_edge_list), len(rsam_edge_attributes))
print('number of nodes', len(rsam_graph_indicator), len(rsam_node_labels), len(rsam_node_attributes))
print('number of graphs', len(rsam_graph_labels), valid_count)


#Load the labels from excel file
# top_classes = ['Sactipeptide', 'Ranthipeptide', 'PQQ', 'Mycofactosin', 'AnSME Sulfatase Filtered']
wb = load_workbook(filename='./data_files/rSAM Enzyme Classes Uniprot IDs_2.xlsx')
top_classes = wb.sheetnames
uid_top_classes = {}
for class_id, sheet_name in enumerate(top_classes):
    ws = wb[sheet_name]
    assert ws.cell(row=1, column=1).value == 'Uniprot IDs'
    count = 0
    row = 2
    break_flag = False
    while not break_flag:
        if ws.cell(row=row, column=1).value is not None:
            count += 1
            uid_top_classes[ws.cell(row=row, column=1).value] = class_id

        # since there can be breaks in the list
        none_count = 0
        for i in range(1, 4):
            if ws.cell(row=row+i, column=1).value is None:
                none_count += 1
        if none_count == 3:
            break_flag = True

        row += 1
    print(sheet_name, count)

print(len(uid_top_classes.keys()))

##################################
##################################
file_prefix = '_activeonlytop5'

fpath = '/data/shared/vishal/rSAM/input_graphs'+file_prefix

folders = os.listdir(fpath)
print(len(folders))

if os.path.isfile('./data_files/residues_list.txt') and os.path.isfile('./data_files/atoms_list.txt'):
    residual_set = []
    with open('./data_files/residues_list.txt', 'r') as f:
        residual_set = f.read().splitlines()
        
    atom_set = []
    with open('./data_files/atoms_list.txt', 'r') as f:
        atom_set = f.read().splitlines()

else:        
    residual_set = set()
    atom_set = set()
    
    for uid in tqdm(folders):
        uid_path = os.path.join(fpath, uid)
        
        if not os.path.isdir(uid_path):
            continue
    
        node_features = np.load(os.path.join(uid_path, 'node_features.npy'))
        
        # read the residual and atom identities
        with open(os.path.join(uid_path, 'node_identities.npy'), 'r') as fr:
            lines = fr.read().splitlines()
    
        assert len(lines) == node_features.shape[0]
        
        for line in lines:
            tmp = line.split('\t')
            residual_set.add(tmp[1])
            atom_set.add(tmp[2])
            
    residual_set = list(residual_set)
    atom_set = list(atom_set)

print(len(residual_set), residual_set)
print(len(atom_set), atom_set)


##################################
##################################


# create mapping for residual set and atom set
residual_mapping, atom_mapping = {}, {}
#residual_set = list(residual_set)
#atom_set = list(atom_set)
for i, residual_name in enumerate(residual_set):
    residual_mapping[residual_name] = i
for i, atom_name in enumerate(atom_set):
    atom_mapping[atom_name] = i

print(residual_mapping)
print(atom_mapping)


folders = set(folders)
uid_top_classes_list = list(uid_top_classes.keys())
valid_count = 0
uid_map = []
num_nodes = 0
num_graphs = 0


rsam_A = {}
rsam_graph_indicator = []
rsam_graph_labels = []
rsam_node_labels = []
rsam_node_attributes = []
rsam_graph_uid = []


for uid in tqdm(uid_top_classes_list):
    if uid not in folders:
        continue
    #print(uid)
    rsam_graph_uid.append(valid_count)
    uid_map.append(uid)
    
    valid_count += 1
    uid_path = os.path.join(fpath, uid)
    # load the processed features
    dist = np.load(os.path.join(uid_path, 'distance_matrix.npy'))
    edge_list = np.load(os.path.join(uid_path, 'edge_list.npy'))
    node_features = np.load(os.path.join(uid_path, 'node_features.npy'))
    tmp_residual_labels, tmp_atom_labels = [], []
    with open(os.path.join(uid_path, 'node_identities.npy'), 'r') as fr:
        lines = fr.read().splitlines()
    assert len(lines) == dist.shape[0] and len(lines) == node_features.shape[0]
    for line in lines:
        line = line.split('\t')
        tmp_residual_labels.append(residual_mapping[line[1]])
        tmp_atom_labels.append(atom_mapping[line[2]])
    
    # first add the edges in edge_list to rsam_A
    for m in range(edge_list.shape[0]):
        # do not include self-loops
        if edge_list[m][0] == edge_list[m][1]:
            continue
        tmp_edge = (num_nodes + edge_list[m][0], num_nodes + edge_list[m][1])
        tmp_edge_rev = (num_nodes + edge_list[m][1], num_nodes + edge_list[m][0])
        assert dist[edge_list[m][0]][edge_list[m][1]] == dist[edge_list[m][1]][edge_list[m][0]]
        rsam_A[tmp_edge] = 1 / dist[edge_list[m][0]][edge_list[m][1]] ** 2
        rsam_A[tmp_edge_rev] = 1 / dist[edge_list[m][0]][edge_list[m][1]] ** 2

    rsam_graph_indicator += [num_graphs] * dist.shape[0]
    rsam_graph_labels.append(int(uid_top_classes[uid]))
    
    rsam_node_labels += tmp_atom_labels
    
    for n in range(dist.shape[0]):
        one_hot_residual_feature = np.zeros(len(residual_set))
        one_hot_residual_feature[tmp_residual_labels[n]] = 1.0
        # one_hot_atom_feature = np.zeros(len(atom_set))
        # one_hot_atom_feature[tmp_atom_labels[n]] = 1.0
        rsam_node_attributes.append(np.concatenate((node_features[n], one_hot_residual_feature)))

    # update the number of graphs and num_nodes
    num_nodes += dist.shape[0]
    num_graphs += 1

# separate rsam_A out as A and edge_attributes
rsam_A_edge_list = []
rsam_edge_attributes = []
for edge in rsam_A:
    rsam_A_edge_list.append([edge[0], edge[1]])
    rsam_edge_attributes.append(rsam_A[edge])

print('number of edges', len(rsam_A_edge_list), len(rsam_edge_attributes))
print('number of nodes', len(rsam_graph_indicator), len(rsam_node_labels), len(rsam_node_attributes))
print('number of graphs', len(rsam_graph_labels), valid_count)


# save data to files
# rsam_top5_A

gnn_files_loc = '/data/shared/vishal/rSAM/gnn_graphs_1/data/rsam'+file_prefix+'/raw/'

with open(os.path.join(gnn_files_loc,'rsam'+file_prefix+'_A.txt'), 'w') as fw:
    for line in rsam_A_edge_list:
        # line = list(map(str, line))
        # fw.write(','.join(line) + '\n')
        fw.write(f'{line[0]+1},{line[1]+1}\n')

# rsam_top5_graph_indicator
with open(os.path.join(gnn_files_loc,'rsam'+file_prefix+'_graph_indicator.txt'), 'w') as fw:
    for tmp in rsam_graph_indicator:
        fw.write(str(tmp+1) + '\n')

# rsam_top5_graph_labels
with open(os.path.join(gnn_files_loc,'rsam'+file_prefix+'_graph_labels.txt'), 'w') as fw:
    for tmp in rsam_graph_labels:
        fw.write(str(tmp) + '\n')
        
with open(os.path.join(gnn_files_loc,'rsam'+file_prefix+'_graph_uid.txt'), 'w') as fw:
    for tmp in rsam_graph_uid:
        fw.write(str(tmp) + '\n')

# rsam_top5_node_labels
with open(os.path.join(gnn_files_loc,'rsam'+file_prefix+'_node_labels.txt'), 'w') as fw:
    for tmp in rsam_node_labels:
        fw.write(str(tmp) + '\n')

# rsam_top5_edge_attributes
with open(os.path.join(gnn_files_loc,'rsam'+file_prefix+'_edge_attributes.txt'), 'w') as fw:
    for tmp in rsam_edge_attributes:
        fw.write(str(tmp) + '\n')

# rsam_top5_node_attributes
with open(os.path.join(gnn_files_loc,'rsam'+file_prefix+'_node_attributes.txt'), 'w') as fw:
    for line in rsam_node_attributes:
        line = list(map(str, line))
        fw.write(','.join(line) + '\n')
        
# rsam_top5_edge_attributes
with open(os.path.join(gnn_files_loc,'rsam'+file_prefix+'_uid_map.txt'), 'w') as fw:
    for tmp in uid_map:
        fw.write(tmp + '\n')
        
'''
        
'''
with open('/data/shared/vishal/rSAM/gnn_graphs_1/data/rsam_activeonlytop5/raw/rsam_activeonlytop5_graph_labels.txt', 'r') as f:
    rsam_uid_labels = f.read().splitlines()
    
with open('/data/shared/vishal/rSAM/gnn_graphs_1/data/rsam_activeonlytop5/raw/rsam_activeonlytop5_graph_labels.txt', 'w') as f:
    for label in rsam_uid_labels:
        if(int(label)>4):
            f.write('5\n')
        else:
            f.write(label)
            f.write('\n')
'''